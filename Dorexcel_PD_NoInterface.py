import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import load_workbook
from datetime import datetime
import xlrd

# Carregar os arquivos de dados
Arq1 = pd.read_excel('Dados/Clientes.xlsx', sheet_name='Customers')
Arq2 = pd.read_excel('Dados/Clientes.xlsx', sheet_name='Addresses')
Arq3 = xlrd.open_workbook("Dados/Vendas_Corrig.xls")

pd.options.mode.chained_assignment = None
# Selecionar apenas as colunas desejadas da Planilha 1
colunas_planilha1 = Arq1[["firstname", "lastname","date_added", "telephone", "email"]]
# Concatenação firstname and lastname
colunas_planilha1["Nome do Cliente"] = colunas_planilha1["firstname"] + " " + colunas_planilha1["lastname"]
# Deletando firstname and lastname da planilha original
colunas_planilha1 = colunas_planilha1.drop(["firstname", "lastname"], axis=1)

#colunas_planilha2.to_excel('Dss.xlsx', index=False)

#print(colunas_planilha2)

# Renomear as colunas da Planilha 1
colunas_planilha1 = colunas_planilha1.rename(columns={"date_added": "Data Cadastrado", "telephone": "Telefone", "email": "Email"})

# Selecionar apenas as colunas desejadas da Planilha 2
colunas_planilha2 = Arq2[["firstname", "lastname", "address_1", "address_2", "city", "postcode", "zone", "custom_field"]]
# Edita a coluna address_1
#colunas_planilha2["address_1"] = colunas_planilha2["address_1"].str.replace("1", "").str.replace("2", "").str.replace("3", "").str.replace("4", "").str.replace("5", "").str.replace("6", "").str.replace("7", "").str.replace("8", "").str.replace("9", "").str.replace(",", "")
# Edita a coluna custom_field
colunas_planilha2["custom_field"] = colunas_planilha2["custom_field"].str.replace("7", "").str.replace("8", "").str.replace("{", "").str.replace("}", "").str.replace(":", "").str.replace(",", ", ").str.replace("\"", "")
# Concatenação firstname and lastname
colunas_planilha2["Nome do Cliente"] = colunas_planilha2["firstname"] + " " + colunas_planilha2["lastname"]
# Concatenação das colunas de endereço solicitadas
colunas_planilha2["Endereço"] = colunas_planilha2["address_1"] + " " + colunas_planilha2["custom_field"] + " - " + colunas_planilha2["address_2"] + " - " + colunas_planilha2["city"] + " - " + colunas_planilha2["zone"] + " - " + colunas_planilha2["postcode"]
# Deletando as colunas que não desejamos para a manipulação
colunas_planilha2 = colunas_planilha2.drop(["firstname", "lastname", "address_1", "address_2", "city", "postcode", "zone", "custom_field"], axis=1)

# Manipulação Arq3
index = Arq3.sheet_by_index(0)
data = []

for rx in range(index.nrows):
    data.append(index.row_values(rx))

col_names = data[0]
data = data[1:]

# Criando Dataframe Arq3
colunas_planilha3 = pd.DataFrame(data, columns=col_names)

# Selecionar apenas as colunas desejadas da Planilha 3
colunas_planilha3 = colunas_planilha3[[ "Email", "Date Added", "Order Status", "Total"]]

# Renomear as colunas da Planilha 3
colunas_planilha3 = colunas_planilha3.rename(columns={"Email": "Email", "Date Added": "Data Última Compra",
                                                    "Order Status": "Total de Compras (Finalizado)", "Total": "Soma dos Valores"})

# Combinação das planilhas usando o email como chave
dados_concatenados = pd.merge(colunas_planilha1, colunas_planilha3, on="Email", how="left")
dados_concatenados = pd.merge(dados_concatenados, colunas_planilha2, on="Nome do Cliente", how="left")

# Adicionar a nova coluna "Link Whatsapp"
dados_concatenados["Link Whatsapp"] = "https://wa.me/55" + dados_concatenados["Telefone"].str.replace("(", "").str.replace(")", "").str.replace("-", "").str.replace(" ", "")

# Converter colunas de data para o tipo datetime
dados_concatenados['Data Cadastrado'] = pd.to_datetime(dados_concatenados['Data Cadastrado']).dt.date
# Formatar a coluna "Data Cadastrado"
#dados_concatenados['Data Cadastrado'] = pd.to_datetime(dados_concatenados['Data Cadastrado']).dt.strftime('%d/%m/%y')

dados_concatenados['Data Última Compra'] = pd.to_datetime(dados_concatenados['Data Última Compra']).dt.date
# Formatar a coluna "Data Última Compra"
#dados_concatenados['Data Última Compra'] = pd.to_datetime(dados_concatenados['Data Última Compra']).dt.strftime('%Y/%m/%d')


# Substituir "Finalizado" por 1 na coluna "Total de Compras (Finalizado)"
dados_concatenados.loc[dados_concatenados['Total de Compras (Finalizado)'] == 'Finalizado', 'Total de Compras (Finalizado)'] = 1

# Remover as linhas em que a coluna "Total de Compras (Finalizado)" contém o valor "cancelado"
dados_concatenados = dados_concatenados[dados_concatenados['Total de Compras (Finalizado)'] != 'Cancelado']

# Converter a coluna "Soma dos Valores" para numérica
dados_concatenados['Soma dos Valores'] = dados_concatenados['Soma dos Valores'].str.replace("R$", "").str.replace(",", "").str.replace(".", "").astype(float)

# Converter a coluna "Total de Compras (Finalizado)" para numérica
dados_concatenados['Total de Compras (Finalizado)'] = pd.to_numeric(dados_concatenados['Total de Compras (Finalizado)'], errors='coerce')

# Somar os valores da coluna "Soma dos Valores" por cliente
dados_somados_soma = dados_concatenados.groupby("Email")["Soma dos Valores"].sum().reset_index()
dados_somados_soma['Soma dos Valores'] = dados_somados_soma['Soma dos Valores'] / 100
# Renomear a coluna somada
dados_somados_soma = dados_somados_soma.rename(columns={"Soma dos Valores": "Soma dos Valores Somada"})

# Somar os valores da coluna "Total de Compras (Finalizado)" por cliente
dados_somados_compras = dados_concatenados.groupby("Email")["Total de Compras (Finalizado)"].sum().reset_index()

# Renomear a coluna somada
dados_somados_compras = dados_somados_compras.rename(columns={"Total de Compras (Finalizado)": "Total de Compras (Finalizado) Somado"})

# Mesclar os dados somados com os dados originais usando a chave "Email"
dados_concatenados = pd.merge(dados_concatenados, dados_somados_soma, on="Email", how="left")
dados_concatenados = pd.merge(dados_concatenados, dados_somados_compras, on="Email", how="left")

# Reordenar as colunas
colunas_desejadas = ["Data Cadastrado", "Data Última Compra", "Nome do Cliente", "Total de Compras (Finalizado) Somado",
                    "Soma dos Valores Somada", "Email", "Telefone", "Link Whatsapp", "Endereço"]

dados_concatenados = dados_concatenados[colunas_desejadas]

dados_concatenados = dados_concatenados.drop_duplicates(subset='Email', keep='first')

dados_concatenados = dados_concatenados.sort_values(by='Soma dos Valores Somada', ascending=False)

# Exibir os dados concatenados
#print(dados_concatenados)

dados_concatenados.to_excel('Dados Site.xlsx', index=False)

# Carregar o arquivo Excel usando openpyxl
wb = load_workbook(filename='Dados Site.xlsx')
ws = wb.active

# Definir a formatação pré-definida
font_bold = Font(bold=True, name='Arial')
font = Font(name='Arial')
alignment_center = Alignment(horizontal="center", vertical="center")

# Ajustar o tamanho das colunas
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Alterar a fonte para Arial
for row in ws.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.font = Font(name='Arial', color="FFFFFF")
        cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

# Centralizar dados nas células e trocar font das cells
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = alignment_center
        cell.font = font

# Verificar a data atual
data_atual = datetime.now().date()

# Percorrer as células da coluna "Data Última Compra"
for row in ws.iter_rows(min_row=2):
    cell_data_ultima_compra = row[1]
    if isinstance(cell_data_ultima_compra.value, datetime):
        if (data_atual - cell_data_ultima_compra.value.date()).days > 30:
            # Pintar a linha inteira com a cor vermelha
            for cell in row:
                cell.fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')

        # Alterar o formato da data para "dd/mm/yyyy"
        cell_data_ultima_compra.number_format = 'dd/mm/yyyy'

for row in ws.iter_rows(min_row=2):
    cell_data_cadastrada = row[0]
        # Alterar o formato da data para "yyyy/mm/dd"
    cell_data_cadastrada.number_format = 'dd/mm/yyyy'

# Definir o estilo de número como moeda
currency_style = 'R$ #,##0.00'

# Aplicar o estilo de número às células da coluna "Soma dos Valores"
for cell in ws['E'][1:]:
    cell.number_format = currency_style

# Adicionar hiperlinks à coluna "Link Whatsapp"
for row in ws.iter_rows(min_row=2):
    cell_link_whatsapp = row[7]
    hyperlink = cell_link_whatsapp.value
    if hyperlink:
        cell_link_whatsapp.hyperlink = hyperlink

# Habilitar classificar e filtrar
ws.auto_filter.ref = ws.dimensions

# Salvar o arquivo com as formatações
wb.save(filename='Dados Site.xlsx')